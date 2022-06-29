import datetime
import re
from numbers import Number

from openpyxl import load_workbook

from db.db_session import init_db, db_session
from db.models import Course
from env import path_to_xldata

FILENAME = path_to_xldata

keys = ['name', 'age_from', 'age_to', 'focus', 'direction', 'description',
        'teachers', 'area', 'free', 'code', 'schedule', 'counter']


def clear(s):
    if s is None:
        return
    return ' '.join(s.split())


def correct_area(area):
    correct = {'Макеева': 'пр. Макеева, 39',
               'Разина': 'ул. Ст. Разина, 4',
               'Первомайская': 'ул. Первомайская, 9',
               'Марта': 'ул. 8-е Марта, 147',
               'Октября': 'пр. Октября, 21'}
    for k in correct:
        if k in area:
            return correct[k]
    return ' '.join(area.split())


def get_age(data):
    if isinstance(data, Number):
        return int(data), int(data)
    if isinstance(data, datetime.datetime):
        return data.day, data.month
    attempt = re.search('(\d*)\+', data)
    if attempt:
        age_from = attempt.groups()[0]
        return int(age_from), 18
    attempt = re.search('(\d*)\s*-\s*(\d*)\s*[\w\W]*', data)
    if attempt:
        age_from, age_to = attempt.groups()
        return int(age_from), int(age_to)


def create_course(name, age, focus, direction, description, teachers, area, free, code, *schedule):
    code = int(code) if code else -1
    if not all((name, age, focus, direction, description, teachers, area, free, code)):
        return None
    schedule = dict(filter(lambda pair: pair[1], zip((str(i) for i in range(1, 11)),
                                                     (clear(value) for value in schedule))))
    if not schedule:
        return None
    name = name.strip()
    age_from, age_to = get_age(age)
    teachers = teachers.split(',')
    focus = focus.upper()
    direction = direction.upper()
    area = correct_area(area)
    free = free is None or free.strip().lower() == 'бюджет'

    course = Course(**dict(zip(keys, (name, age_from, age_to, focus, direction, description, teachers,
                                      area, free, code, schedule))))
    db_session.add(course)


init_db()
wb = load_workbook(FILENAME)
ws = wb.active
for row in filter(lambda r: r[0].value, ws.iter_rows(min_row=5)):
    if len(list(filter(lambda c: c.value, row))) <= 1:
        continue
    create_course(*(cell.value for cell in row))
db_session.commit()
